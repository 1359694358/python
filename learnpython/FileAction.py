'''file mode   r  read mode,w   write  ,if file not exist,will create,
r+ mean  not only can read ,but also can write
a  is append mode
    '''
# f=open("./pytxt.txt","w")
import json

f=open("./pytxt.txt","r+")# r+ mean  not only can read ,but also can write
f.write("this is a file,with use open method")
f.close()

''' witch   can auto close  open close file'''

with open("./withopenfile.txt","w+") as ff:
    ff.write("this this is use with open file")

'''xlrd  xlwt'''
import xlrd
import xlwt
import openpyxl
'''write xls'''
excel=xlwt.Workbook()# create excel obj
sheet=excel.add_sheet("python_sheet") # add a sheet with name
# sheet.write(0,0,"hello")  # write  row  col  value
listArr=[[1,2,3,4,5,6],[7,8,9,10,11,12]]
for index in range(len(listArr)):
    for rowD in  range(len(listArr[index])):
        sheet.write(index, rowD, str(listArr[index][rowD]))
excel.save("./python_handle_excel.xls")
print("================================")
#read excel
readExce=xlrd.open_workbook("./python_handle_excel.xls")
# readExce=openpyxl.load_workbook("./python_handle_excel.xls")
# readSheet=readExce.sheet_by_index(0)
readSheet=readExce.sheet_by_name("python_sheet")
rows=readSheet.nrows
cols=readSheet.ncols
for rowIndex in range(rows):
    print(readSheet.row_values(rowIndex,0,cols))
    print("-----")
    for colIndex in range(cols):
        print(readSheet.col_values(colIndex))
        print("||||||")
        print(readSheet.cell_value(rowIndex,colIndex))



#  openpyxl  not cover , and index start from 1;

f=open("./openpyxlread.txt","r+")
result=f.readlines()
jsonData=""
for itemJson in result:
    jsonData=jsonData+itemJson
caseJson=json.loads(jsonData)
print(caseJson)
wb=openpyxl.Workbook()#direct create a xlsx file
wb=openpyxl.load_workbook("./openpy.xlsx") # load an exist file
print(wb.sheetnames)
wb._sheets
pysheet=wb.active
pysheet.title=u"pyxls title"
appendTitle=False
for index in caseJson:
    keys= []
    values= []
    for keyItem in dict(index).keys():
        keys.append(keyItem)

    for valueItem in dict(index).values():
        values.append(valueItem)
    print((keys))
    print((values))
    if appendTitle==False:
        pysheet.append(keys)
        appendTitle=True
    pysheet.append(values)
wb.save("./openpy.xlsx")
#
# readExce=openpyxl.load_workbook("./python_handle_excel.xls")
# readSheet=readExce["python_sheet"]
# rows=readSheet.nrows
# cols=readSheet.ncols
# for rowIndex in range(rows):
#     print(readSheet.row_values(rowIndex,0,cols))
#     print("-----")
#     for colIndex in range(cols):
#         print(readSheet.col_values(colIndex))
#         print("||||||")
#         print(readSheet.cell_value(rowIndex,colIndex))
